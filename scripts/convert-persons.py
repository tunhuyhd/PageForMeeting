"""Convert the two class attendance workbooks to import-ready person JSON files."""

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCES = (
    (ROOT / "public" / "Q1 tham gia 15-8.xlsx", "96Q1", ROOT / "public" / "persons-96q1.json"),
    (ROOT / "public" / "Q2 tham gia 15-8 2.xlsx", "96Q2", ROOT / "public" / "persons-96q2.json"),
)


def clean_name(family_name, given_name):
    family_name = " ".join(str(family_name).split())
    given_name = " ".join(str(given_name).split())
    # Some source rows repeat the given name at the end of the family-name cell.
    if family_name.casefold().split()[-1] == given_name.casefold():
        return family_name
    return f"{family_name} {given_name}"


def convert(source, class_name):
    sheet = load_workbook(source, read_only=True, data_only=True).active
    people = []
    for row in sheet.iter_rows(values_only=True):
        sequence, family_name, given_name = row[:3]
        is_person_row = isinstance(sequence, (int, float)) or (
            isinstance(sequence, str) and sequence.strip().isdigit()
        )
        if not is_person_row or not family_name or not given_name:
            continue

        full_name = clean_name(family_name, given_name)
        source_note = str(row[10] or "").strip() if len(row) > 10 else ""
        people.append(
            {
                "fullName": full_name,
                "className": class_name,
                "isFemale": "th\u1ecb" in full_name.casefold().split(),
                "isAttending": True,
                "phone": "",
                "note": source_note,
                "roomKey": None,
                "hotelId": None,
                "hotelName": None,
                "roomNumber": None,
                "dinnerTableId": None,
                "dinnerTableName": None,
            }
        )
    return people


for source_path, person_class, output_path in SOURCES:
    converted = convert(source_path, person_class)
    output_path.write_text(
        json.dumps(converted, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"{output_path.name}: {len(converted)} persons")
